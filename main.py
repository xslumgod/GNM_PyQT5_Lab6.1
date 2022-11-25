import os
import sys
from pathlib import Path
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
from docxtpl import DocxTemplate
from openpyxl import load_workbook

class Main(QDialog):
    def __init__(self):
        super(Main, self).__init__()
        loadUi('uis/main.ui', self)
        self.setWindowIcon(QtGui.QIcon('images/icon.png'))
        self.setWindowTitle('Создание копий визитки в Word и Excel с помощью Python')

        self.btn_wrd.clicked.connect(self.execute_wrd)
        self.btn_xl.clicked.connect(self.execute_xl)

    def execute_wrd(self):
        document_path = Path(__file__).parent / "template.docx"
        doc = DocxTemplate(document_path)
        context = {"Corp": self.lineEdit.text(),
                   "First_last_name": self.lineEdit_2.text(),
                   "title": self.lineEdit_3.text(),
                   "Mail": self.lineEdit_4.text(),
                   "Phone": self.lineEdit_5.text(),
                   "Street": self.lineEdit_6.text(),
                   "City": self.lineEdit_7.text(),
                   "House": self.lineEdit_9.text(),
                   "website": self.lineEdit_8.text()}
        doc.render(context)
        doc.save(Path(__file__).parent / "generatedd.docx")
        os.system('start generatedd.docx')

    def execute_xl(self):
        fn = 'template.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['A1'] = self.lineEdit.text()
        ws['B3'] = self.lineEdit_2.text()
        ws['B4'] = self.lineEdit_3.text()
        ws['B6'] = self.lineEdit_4.text()
        ws['B8'] = self.lineEdit_5.text()
        ws['B9'] = self.lineEdit_5.text()
        ws['B10'] = self.lineEdit_5.text()
        ws['E7'] = self.lineEdit_9.text()
        ws['E8'] = self.lineEdit_6.text()
        ws['E9'] = self.lineEdit_7.text()
        ws['E10'] = self.lineEdit_8.text()

        wb.save(Path(__file__).parent / "generated.xlsx")
        wb.close()
        os.system('start generated.xlsx')

def main():
    app = QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
